import glob
import pandas as pd
import pandasql
import numpy as np
import xlsxwriter
from openpyxl import load_workbook
FilePath = r'inputforgen.xlsx'
ExcelWorkbook = load_workbook(FilePath)
ws = ExcelWorkbook.active

xval_list = []
for x in list(ws.columns)[0]:
        xval_list.append(x.value)
print(xval_list)

yval_list = []
for y in list(ws.columns)[1]:
        yval_list.append(y.value)
        
print(yval_list)       
zval_list = []
for z in list(ws.columns)[2]:
        zval_list.append(z.value)
print(zval_list) 

agg_list = []
for a in list(ws.columns)[3]:
        agg_list.append(a.value)
print(agg_list) 
        
def schema_categorise(row):  
    if row['Table1_Column'] == row['Table2_Column'] :
        return 'Matches'
    return 'NotMatches'

def categorise(row,Grand1_Total,Grand2_Total):  
    if row['Grand1_Total'] == row['Grand2_Total'] :
        return 'Matches'
    return 'NotMatches'
pd.set_option('display.max_columns', None)
df1 = pd.read_csv(r'C:\auto_test\table1.csv',sep=',', header='infer', index_col =0)
df2 = pd.read_csv(r'C:\auto_test\table2.csv',sep=',', header='infer', index_col =0)


outPath = r'C:\auto_test\output.xlsx'
ExcelWorkbook = load_workbook(outPath)
writer = pd.ExcelWriter(outPath, engine = 'openpyxl')
writer.book = ExcelWorkbook

#Schema_Validation
collis1 = (df1.columns).T
collis2 = (df2.columns).T
l1 = pd.DataFrame(collis1, columns = ['Table1_Column'])
l2 = pd.DataFrame(collis2, columns = ['Table2_Column'])
resultschema = pd.DataFrame.merge(l1, l2, left_on = 'Table1_Column', right_on = 'Table2_Column' , how='left')
resultschema['compare'] = resultschema.apply(lambda row: schema_categorise(row), axis=1)
resultstyle = resultschema.style.apply(lambda x: ['background-color:red' if x == 'NotMatches' else 'background:green' for x in resultschema.compare ], axis = 0)
resultstyle.to_excel(writer, sheet_name='schema validation',engine ='openpyxl',header=True, index = False)


for X in range(1,len(yval_list)):
    dfz1 = df1.pivot_table(index = yval_list[X], columns = xval_list[X] ,values= zval_list[X], margins = True, margins_name = 'Grand1_Total' ,aggfunc= agg_list[X])
    dfz2 = df2.pivot_table(index = yval_list[X], columns = xval_list[X] ,values= zval_list[X], margins = True, margins_name = 'Grand2_Total' ,aggfunc= agg_list[X])
    resultdf = pd.DataFrame.merge(dfz1, dfz2, left_on = yval_list[X], right_on = yval_list[X] , how='left')
    resultdf['compare'] = resultdf.apply(lambda row: categorise(row, 'Grand1_Total','Grand2_Total' ), axis=1)
    resultdf['variance'] = ((resultdf['Grand1_Total'] - resultdf['Grand2_Total'])/resultdf['Grand1_Total'])*100
    resultstyledf = resultdf.style.apply(lambda x: ['background-color:red' if x == 'NotMatches' else 'background:green' for x in resultdf.compare ], axis = 0)
    resultstyledf.to_excel(writer, sheet_name= zval_list[X],engine ='openpyxl', header=True, index = True)
  
    
writer.save()
writer.close()